from PyPDF2 import PdfReader
import spacy
from pathlib import Path
from os import path
from openpyxl import Workbook
import customtkinter
from plyer import notification
from tkinter.filedialog import askdirectory
import threading

# Modes: "System" (standard), "Dark", "Light"
customtkinter.set_appearance_mode("dark")
# Themes: "blue" (standard), "green", "dark-blue"
customtkinter.set_default_color_theme("blue")

app = customtkinter.CTk()
app.geometry("400x780")
app.title("Invoixe")

main_frame = customtkinter.CTkFrame(master=app)
main_frame.pack(pady=20, padx=60, fill="both", expand=True)

header_label = customtkinter.CTkLabel(
    master=main_frame, justify=customtkinter.LEFT, text="Invoixe")
header_label.pack(pady=10, padx=10)

files_progress = customtkinter.CTkProgressBar(master=main_frame)
files_progress.pack(pady=10, padx=10)
files_progress.set(0)

invoice_ent_names = {"GPE": "Location",
                     "DATE": "Date", "ORG": "Company", "MONEY": "Total"}


def get_invoice_info(file_path) -> dict:
    try:
        reader = PdfReader(file_path)
        text = reader.pages[0].extract_text()

        nlp = spacy.load('./nlp_serialized')

        doc = nlp(text)
        doc_info = dict()
        for ent in doc.ents:
            if ent.label_ in invoice_ent_names.keys():
                doc_info[ent.label_] = ent.text

        return doc_info
    except Exception as e:
        print(f"Failed to parse file {file_path}, Error: {e}")

        return dict()


def create_execl_of_invoices(files, save_loc):
    wb = Workbook()
    sheet_name = "Sheet1"
    wb.create_sheet(sheet_name)
    work_sheet = wb[sheet_name]

    header_col = 1
    for value in invoice_ent_names.values():
        work_sheet.cell(1, header_col, value)
        header_col += 1
    work_sheet.cell(1, header_col, "File")

    progress_step = 100 / len(files)
    progress = 0

    row_start = 1  # start below the header row 1
    for f in files:
        doc_info = get_invoice_info(f)

        # Nothing found
        if len(doc_info) == 0:
            continue

        row_start += 1
        col_start = 1  # starts from column B
        for key in invoice_ent_names.keys():
            work_sheet.cell(row_start, col_start, doc_info.get(key))
            col_start += 1

        work_sheet.cell(row_start, col_start,
                        path.splitext(path.basename(f))[0])

        files_progress.set(progress)
        progress += progress_step

    wb.save(save_loc)
    wb.close()


def handle_dir_processing():
    selected_dir = askdirectory(
        initialdir=path.join(Path.home(), "Downloads"))

    files = list(Path(selected_dir).rglob('*.pdf'))

    create_execl_of_invoices(files, path.join(
        selected_dir, "./results.xlsx"))

    notification.notify(
        title="Invoixe", message="All files processed successfully")


def handle_dir_select():
    t = threading.Thread(target=handle_dir_processing)
    t.start()


def slider_callback(value):
    files_progress.set(value)


select_dirs_btn = customtkinter.CTkButton(
    master=main_frame, command=handle_dir_select, text="Select directory on invoices")
select_dirs_btn.pack(pady=10, padx=10)

app.mainloop()
